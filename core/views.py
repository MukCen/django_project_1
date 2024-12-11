from django.shortcuts import get_object_or_404, redirect, render
from django.http import Http404, JsonResponse, HttpResponse

from core.forms import PaymentForm
from .models import Products, Client, Order, Payment
import openpyxl
from django.utils.encoding import smart_str
from django.core.files.storage import FileSystemStorage
from django.conf import settings
import os
from django.views.generic import DetailView, ListView

from core.utils import q_search


def product_list(request):
    products = Products.objects.all()
    return render(request, 'core/product_list.html', {'products': products})


def export_to_excel(request):
    # Create a workbook and add a worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Products"

    # Define the titles for columns
    columns = ['Name', 'Description', 'Price', 'Quantity']
    ws.append(columns)

    # Fetch data from the database and append it to the worksheet
    products = Products.objects.all()
    for product in products:
        ws.append([product.name, product.description, product.price, product.quantity])

    # Prepare the response and save the file in memory
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=products.xlsx'
    wb.save(response)  # type: ignore

    return response


def import_from_excel(request):
    if request.method == 'POST' and request.FILES['import_file']:
        import_file = request.FILES['import_file']
        fs = FileSystemStorage(location=os.path.join(settings.MEDIA_ROOT, 'uploads'))
        filename = fs.save(import_file.name, import_file)
        filepath = fs.path(filename)

        wb = openpyxl.load_workbook(filepath)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):  # Skip the header row
            name, description, price, quantity = row
            Products.objects.create(
                name=name, description=description, price=price, quantity=quantity
            )

        # Delete the file after processing
        os.remove(filepath)

        return JsonResponse({'status': 'success'})

    return JsonResponse({'status': 'fail', 'message': 'No file uploaded'})


def client_list(request):
    clients = Client.objects.all()
    return render(request, 'core/client_list.html', {'clients': clients})


def order_list(request):
    orders = Order.objects.all()
    return render(request, 'core/order_list.html', {'orders': orders})


def payment_list(request):
    payments = Payment.objects.select_related('order').all()
    return render(request, 'core/payment_list.html', {'payments': payments})


def add_payment(request):
    if request.method == 'POST':
        form = PaymentForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect(
                'payment_success'
            )  # Направляє на сторінку успішного додавання
    else:
        form = PaymentForm()
    return render(request, 'core/add_payment.html', {'form': form})


def order_details(request, order_id):
    # Отримання замовлення за його ID
    order = get_object_or_404(Order, id=order_id)

    # Отримання інформації про оплату (якщо є)
    payment = Payment.objects.filter(order=order).first()

    context = {
        'order': order,
        'payment': payment,
    }
    return render(request, 'core/order_details.html', context)


class CatalogView(ListView):
    model = Products
    # queryset = Products.objects.all().order_by("-id")
    template_name = "goods/catalog.html"
    context_object_name = "goods"
    paginate_by = 3
    allow_empty = False
    # чтоб удобно передать в методы
    slug_url_kwarg = "category_slug"

    def get_queryset(self):
        category_slug = self.kwargs.get(self.slug_url_kwarg)
        on_sale = self.request.GET.get("on_sale")
        order_by = self.request.GET.get("order_by")
        query = self.request.GET.get("q")

        if category_slug == "all":
            goods = super().get_queryset()
        elif query:
            goods = q_search(query)
        else:
            goods = super().get_queryset().filter(category__slug=category_slug)
            if not goods.exists():
                raise Http404()

        if on_sale:
            goods = goods.filter(discount__gt=0)

        if order_by and order_by != "default":
            goods = goods.order_by(order_by)

        return goods

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "Home - Каталог"
        context["slug_url"] = self.kwargs.get(self.slug_url_kwarg)
        return context


class ProductView(DetailView):

    # model = Products
    # slug_field = "slug"
    template_name = "core/product.html"
    slug_url_kwarg = "product_slug"
    context_object_name = "product"

    def get_object(self, queryset=None):
        product = Products.objects.get(slug=self.kwargs.get(self.slug_url_kwarg))
        return product

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = self.object.name  # type: ignore
        return context


# def payment_list(request):
#     payments = Payment.objects.all()
#     return render(request, 'core/payment_list.html', {'payments': payments})
